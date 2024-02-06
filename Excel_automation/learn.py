import openpyxl

# wb = workbook
# ws = worksheet

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("file1.xlsx")

ws = wb.worksheets[0]

# for i in range(1, 10):
#     for j in range(1, 10):
#         ws.cell(row=j, column=i, value=f'{i}x{j}={i*j:2}')

rows = ws['A':'I']

for row in rows:
    for col in row:
        print(col.value)

# wb.save("test.xlsx")
